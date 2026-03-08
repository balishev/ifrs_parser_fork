from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from .metrics import load_metrics
from .parser import GoogleIFRSPdfParser, IFRSParserConfig

REQUIRED_METRICS = (
    "revenue",
    "interest_expense_loans",
    "depreciation",
    "cash_and_cash_equivalents",
    "property_plant_and_equipment",
    "operating_profit",
    "long_term_debt_and_lease",
    "short_term_debt_and_lease",
)

ROW_TO_METRIC_KEY = {
    "выручка": "revenue",
    "расходы": "interest_expense_loans",
    "%расходы": "interest_expense_loans",
    "процентныерасходы": "interest_expense_loans",
    "амортизация": "depreciation",
    "денежныесредстваиэкв": "cash_and_cash_equivalents",
    "ос": "property_plant_and_equipment",
    "операционнаяприбыль": "operating_profit",
}

COL_LTM = 10
COL_YEAR = {2025: 12, 2024: 13, 2023: 14, 2022: 15}
DEBT_LONG_KEY = "long_term_debt_and_lease"
DEBT_SHORT_KEY = "short_term_debt_and_lease"

DEFAULT_ALIAS_MAP = {
    "вымпелком": "ВЫМПЕЛКОМ",
    "вымпелкоммуникации": "ВЫМПЕЛКОМ",
    "лукойл": "ЛУКОЙЛ",
    "новатэк": "НОВАТЭК",
    "сибур": "СИБУР",
    "мираторг": "МИРАТОРГ",
    "цаэк": "ЦАЭК",
    "транскомплектхолдинг": "ТРАНСКОМПЛЕКТХОЛДИНГ",
    "трансмашхолдинг": "ТРАНСМАШХОЛДИНГ",
    "овк": "ОВК",
    "ммк": "МЕТАЛЛУРГИЯ",
    "magnitogorskironandsteelworks": "МЕТАЛЛУРГИЯ",
    "magnitogorskironsteelworks": "МЕТАЛЛУРГИЯ",
    "magnitogorskmetallurgicalcombine": "МЕТАЛЛУРГИЯ",
}

EXCLUDED_SHEETS = {
    "рашников",
    "линники",
    "мордашов",
    "лисин",
    "усманов",
    "махмудов",
}


@dataclass(slots=True)
class MetricRef:
    ltm: float | None = None
    year_values: dict[int, float] | None = None
    unit: str | None = None
    source_sheet: str | None = None


@dataclass(slots=True)
class CompanyRef:
    company: str
    metrics: dict[str, MetricRef]
    sheets: set[str]


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="ifrs-compare-excel",
        description="Parse IFRS PDFs and compare with one-pager Excel values.",
    )
    parser.add_argument("--root", required=True, help="Root folder with nested .../мсфо/*.pdf files.")
    parser.add_argument("--xlsx", required=True, help="Path to consolidated Excel file.")
    parser.add_argument("--out-csv", default="output/comparison_report.csv", help="Path to output CSV report.")
    parser.add_argument(
        "--out-summary-json",
        default="output/comparison_summary.json",
        help="Path to output JSON summary.",
    )
    parser.add_argument(
        "--parse-cache-dir",
        default="output/parsed_pdfs",
        help="Directory for per-PDF parsed JSON cache.",
    )
    parser.add_argument("--credentials-json", required=True, help="Service account JSON for Vertex AI mode.")
    parser.add_argument("--project", required=True, help="Google Cloud project id.")
    parser.add_argument("--location", default="us-central1", help="Vertex location.")
    parser.add_argument("--model", default="gemini-2.5-flash", help="Model name.")
    parser.add_argument(
        "--fallback-model",
        default="gemini-2.0-flash",
        help="Optional fallback model if primary model call fails.",
    )
    parser.add_argument("--timeout-sec", type=int, default=300, help="PDF processing timeout.")
    parser.add_argument(
        "--include-banks",
        action="store_true",
        help="Include bank reports. By default bank reports are skipped.",
    )
    parser.add_argument(
        "--reparse",
        action="store_true",
        help="Ignore cache and re-run model parsing for all PDFs.",
    )
    return parser


def main() -> int:
    args = build_arg_parser().parse_args()

    root = Path(args.root).resolve()
    xlsx_path = Path(args.xlsx).resolve()
    out_csv = Path(args.out_csv).resolve()
    out_summary = Path(args.out_summary_json).resolve()
    cache_dir = Path(args.parse_cache_dir).resolve()
    cache_dir.mkdir(parents=True, exist_ok=True)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    out_summary.parent.mkdir(parents=True, exist_ok=True)

    company_refs = load_company_refs(xlsx_path)
    all_pdf_paths = sorted(find_ifrs_pdfs(root))
    skipped_bank_pdfs: list[Path] = []
    if not args.include_banks:
        pdf_paths = []
        for path in all_pdf_paths:
            if is_bank_pdf(path):
                skipped_bank_pdfs.append(path)
            else:
                pdf_paths.append(path)
    else:
        pdf_paths = all_pdf_paths

    primary_parser = GoogleIFRSPdfParser(
        credentials_json=args.credentials_json,
        project=args.project,
        config=IFRSParserConfig(
            model=args.model,
            location=args.location,
            timeout_sec=args.timeout_sec,
        ),
    )
    fallback_parser: GoogleIFRSPdfParser | None = None
    if args.fallback_model and args.fallback_model != args.model:
        fallback_parser = GoogleIFRSPdfParser(
            credentials_json=args.credentials_json,
            project=args.project,
            config=IFRSParserConfig(
                model=args.fallback_model,
                location=args.location,
                timeout_sec=args.timeout_sec,
            ),
        )
    metrics = load_metrics()

    rows: list[dict[str, Any]] = []
    parsed_ok = 0
    parse_failed = 0
    matched_company_count = 0
    unmatched_company_count = 0

    for pdf_path in pdf_paths:
        relative_pdf = pdf_path.relative_to(root)
        cache_file = cache_dir / (sanitize_filename(str(relative_pdf)) + ".json")
        parse_payload: dict[str, Any] | None = None
        parse_error: str | None = None

        if cache_file.exists() and not args.reparse:
            parse_payload = json.loads(cache_file.read_text(encoding="utf-8"))
        else:
            try:
                parse_payload = primary_parser.extract_metrics(pdf_path=pdf_path, metrics=metrics)
            except Exception as exc:
                parse_error = str(exc)
                if fallback_parser is not None:
                    try:
                        parse_payload = fallback_parser.extract_metrics(pdf_path=pdf_path, metrics=metrics)
                        parse_error = None
                    except Exception as fallback_exc:
                        parse_error = f"Primary error: {exc}; Fallback error: {fallback_exc}"

            if parse_payload is not None:
                cache_file.write_text(json.dumps(parse_payload, ensure_ascii=False, indent=2), encoding="utf-8")

        if parse_payload is None:
            parse_failed += 1
            rows.append(
                {
                    "pdf_path": str(relative_pdf),
                    "status": "parse_error",
                    "message": parse_error,
                }
            )
            continue

        parsed_ok += 1
        company_match = match_company(parse_payload, relative_pdf, company_refs)
        if company_match is not None and company_match.upper() in company_refs:
            matched_company_count += 1
        else:
            unmatched_company_count += 1

        parsed_period_end = parse_payload.get("reporting_period_end_date")
        parsed_year = extract_year(parsed_period_end)
        parsed_currency = parse_payload.get("reporting_currency")

        for metric in parse_payload.get("metrics", []):
            if metric.get("metric_key") not in REQUIRED_METRICS:
                continue

            row = compare_metric(
                relative_pdf=relative_pdf,
                parse_payload=parse_payload,
                metric_payload=metric,
                parsed_year=parsed_year,
                parsed_currency=parsed_currency,
                company_match=company_match,
                company_refs=company_refs,
            )
            rows.append(row)

    write_csv(rows, out_csv)
    summary = {
        "root": str(root),
        "xlsx": str(xlsx_path),
        "total_pdfs_found": len(all_pdf_paths),
        "total_pdfs_processed": len(pdf_paths),
        "skipped_banks_count": len(skipped_bank_pdfs),
        "parsed_ok": parsed_ok,
        "parse_failed": parse_failed,
        "matched_company_count": matched_company_count,
        "unmatched_company_count": unmatched_company_count,
        "output_csv": str(out_csv),
        "cache_dir": str(cache_dir),
    }
    out_summary.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def find_ifrs_pdfs(root: Path) -> list[Path]:
    result: list[Path] = []
    for path in root.rglob("*.pdf"):
        parts_lower = [part.lower() for part in path.parts]
        if "мсфо" in parts_lower:
            result.append(path)
    return result


def is_bank_pdf(path: Path) -> bool:
    text = normalize_text(str(path))
    bank_markers = (
        "тбанк",
        "tbank",
        "росбанк",
        "rosbank",
        "альфабанк",
        "alfabank",
        "alphabank",
        "альфабанк",
    )
    return any(marker in text for marker in bank_markers)


def load_company_refs(xlsx_path: Path) -> dict[str, CompanyRef]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    refs: dict[str, CompanyRef] = {}
    debt_rows: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        if normalize_sheet_name(ws.title) in EXCLUDED_SHEETS:
            continue
        for r in range(2, ws.max_row + 1):
            company = as_str(ws.cell(r, 2).value)
            if not company:
                continue

            source = as_str(ws.cell(r, 8).value)
            if not source or "мсфо" not in source.lower():
                continue

            metric_name = normalize_metric_name(ws.cell(r, 5).value)
            seg6 = normalize_text(ws.cell(r, 6).value)
            seg7 = normalize_text(ws.cell(r, 7).value)
            unit = as_str(ws.cell(r, 9).value)
            company_upper = company.upper()

            ref = refs.setdefault(
                company_upper,
                CompanyRef(company=company, metrics={}, sheets=set()),
            )
            ref.sheets.add(ws.title)

            metric_key = ROW_TO_METRIC_KEY.get(metric_name)
            if metric_key:
                metric_ref = ref.metrics.setdefault(
                    metric_key,
                    MetricRef(ltm=None, year_values={}, unit=unit, source_sheet=ws.title),
                )
                metric_ref.ltm = to_float(ws.cell(r, COL_LTM).value, metric_ref.ltm)
                for year, col in COL_YEAR.items():
                    year_value = to_float(ws.cell(r, col).value, None)
                    if year_value is not None:
                        metric_ref.year_values[year] = year_value
                if not metric_ref.unit:
                    metric_ref.unit = unit
                continue

            if metric_name != "долг":
                continue

            debt_type = debt_type_from_horizon(seg7)
            if debt_type is None:
                continue

            component_kind = debt_component_kind(seg6)
            if component_kind is None:
                continue

            debt_rows.append(
                {
                    "company_upper": company_upper,
                    "debt_type": debt_type,
                    "component_kind": component_kind,
                    "sheet": ws.title,
                    "unit": unit,
                    "ltm": to_float(ws.cell(r, COL_LTM).value, 0.0) or 0.0,
                    "year_values": {
                        year: to_float(ws.cell(r, col).value, 0.0) or 0.0
                        for year, col in COL_YEAR.items()
                    },
                }
            )

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in debt_rows:
        grouped[(row["company_upper"], row["debt_type"])].append(row)

    for (company_upper, debt_type), items in grouped.items():
        has_lease = any(item["component_kind"] == "lease" for item in items)
        selected_items = [
            item
            for item in items
            if item["component_kind"] in {"credit", "lease"}
            or (item["component_kind"] == "other" and not has_lease)
        ]
        if not selected_items:
            continue

        ltm_total = sum(item["ltm"] for item in selected_items)
        year_totals = {year: 0.0 for year in COL_YEAR}
        for item in selected_items:
            for year, value in item["year_values"].items():
                year_totals[year] += value

        ref = refs.get(company_upper)
        if ref is None:
            continue

        metric_ref = ref.metrics.setdefault(
            debt_type,
            MetricRef(ltm=0.0, year_values={}, unit=None, source_sheet=None),
        )
        metric_ref.ltm = round(ltm_total, 6)
        metric_ref.year_values = {year: round(value, 6) for year, value in year_totals.items()}
        if not metric_ref.unit:
            metric_ref.unit = next((item["unit"] for item in selected_items if item["unit"]), None)
        if not metric_ref.source_sheet:
            metric_ref.source_sheet = selected_items[0]["sheet"]

    return refs


def compare_metric(
    relative_pdf: Path,
    parse_payload: dict[str, Any],
    metric_payload: dict[str, Any],
    parsed_year: int | None,
    parsed_currency: str | None,
    company_match: str | None,
    company_refs: dict[str, CompanyRef],
) -> dict[str, Any]:
    metric_key = metric_payload.get("metric_key")
    parsed_found = bool(metric_payload.get("found"))
    parsed_value = as_float(metric_payload.get("value"))

    base = {
        "pdf_path": str(relative_pdf),
        "company_from_model": as_str(parse_payload.get("company_name")),
        "company_matched": company_match,
        "reporting_period": as_str(parse_payload.get("reporting_period")),
        "reporting_period_end_date": as_str(parse_payload.get("reporting_period_end_date")),
        "reporting_currency": parsed_currency,
        "metric_key": metric_key,
        "metric_found": parsed_found,
        "parsed_value_bn_rub": parsed_value,
        "parsed_notes": as_str(metric_payload.get("notes")),
        "selection_level": as_str(metric_payload.get("selection_level")),
        "source_page": metric_payload.get("page"),
        "source_evidence": as_str(metric_payload.get("evidence")),
    }

    if not company_match:
        base.update(
            {
                "status": "company_not_matched_in_excel",
                "expected_ltm": None,
                "expected_year": None,
                "expected_year_label": None,
                "reference_unit": None,
                "diff_selected": None,
                "diff_pct_selected": None,
            }
        )
        return base

    company_ref = company_refs.get(company_match.upper())
    if not company_ref:
        base.update(
            {
                "status": "company_not_found_in_excel",
                "expected_ltm": None,
                "expected_year": None,
                "expected_year_label": None,
                "reference_unit": None,
                "diff_selected": None,
                "diff_pct_selected": None,
            }
        )
        return base

    metric_ref = company_ref.metrics.get(metric_key)
    if not metric_ref:
        base.update(
            {
                "status": "metric_not_in_excel",
                "expected_ltm": None,
                "expected_year": None,
                "expected_year_label": None,
                "reference_unit": None,
                "diff_selected": None,
                "diff_pct_selected": None,
            }
        )
        return base

    expected_ltm = metric_ref.ltm
    expected_year = metric_ref.year_values.get(parsed_year) if parsed_year and metric_ref.year_values else None
    if expected_year is not None and abs(expected_year) > 1e-12:
        selected_value = expected_year
        selected_label = str(parsed_year)
    elif expected_ltm is not None:
        selected_value = expected_ltm
        selected_label = "LTM"
    else:
        selected_value = expected_year
        selected_label = str(parsed_year) if expected_year is not None and parsed_year is not None else "LTM"
    unit = as_str(metric_ref.unit)

    diff_selected = None
    diff_pct_selected = None
    status = "ok"
    if unit and "руб" not in unit.lower():
        status = "reference_not_rub"
    elif not parsed_found or parsed_value is None:
        status = "metric_not_found_by_parser"
    elif selected_value is None:
        status = "expected_missing"
    else:
        diff_selected = round(parsed_value - selected_value, 6)
        if selected_value != 0:
            diff_pct_selected = round((diff_selected / selected_value) * 100.0, 4)

    base.update(
        {
            "status": status,
            "expected_ltm": expected_ltm,
            "expected_year": expected_year,
            "expected_year_label": selected_label,
            "reference_unit": unit,
            "diff_selected": diff_selected,
            "diff_pct_selected": diff_pct_selected,
            "source_sheet": metric_ref.source_sheet,
        }
    )
    return base


def match_company(
    parse_payload: dict[str, Any],
    relative_pdf: Path,
    company_refs: dict[str, CompanyRef],
) -> str | None:
    candidates = []
    company_name = as_str(parse_payload.get("company_name"))
    if company_name:
        candidates.append(company_name)

    lower_parts = [part.lower() for part in relative_pdf.parts]
    for idx, part in enumerate(lower_parts):
        if part == "мсфо":
            if idx > 0:
                candidates.append(relative_pdf.parts[idx - 1])
            if idx > 1:
                candidates.append(relative_pdf.parts[idx - 2])
    candidates.append(relative_pdf.stem)

    ref_norm_map = {normalize_text(ref.company): ref.company.upper() for ref in company_refs.values()}
    best_company: str | None = None
    best_score = 0

    for candidate in candidates:
        norm_candidate = normalize_text(candidate)
        if not norm_candidate:
            continue

        alias = DEFAULT_ALIAS_MAP.get(norm_candidate)
        if alias:
            return alias

        for norm_ref, ref_company in ref_norm_map.items():
            score = similarity_score(norm_candidate, norm_ref)
            if score > best_score:
                best_score = score
                best_company = ref_company

    if best_score >= 5:
        return best_company
    return None


def similarity_score(a: str, b: str) -> int:
    if not a or not b:
        return 0
    if a == b:
        return 100
    if a in b:
        return min(len(a), len(b))
    if b in a:
        return min(len(a), len(b))
    return common_substring_score(a, b)


def common_substring_score(a: str, b: str) -> int:
    max_len = 0
    for i in range(len(a)):
        for j in range(i + 1, len(a) + 1):
            chunk = a[i:j]
            if len(chunk) <= max_len:
                continue
            if chunk in b:
                max_len = len(chunk)
    return max_len


def debt_type_from_horizon(seg7: str) -> str | None:
    if seg7 not in {"дс", "кс"}:
        return None
    return DEBT_LONG_KEY if seg7 == "дс" else DEBT_SHORT_KEY


def debt_component_kind(seg6: str) -> str | None:
    if not seg6:
        return None
    if seg6 in {"всего", "-", "–"}:
        return None
    if "лизинг" in seg6 or "аренд" in seg6:
        return "lease"
    credit_keywords = (
        "кредит",
        "займ",
        "loan",
        "borrow",
        "bond",
        "облигац",
        "финансов",
    )
    if any(word in seg6 for word in credit_keywords):
        return "credit"
    if "прочие" in seg6 or "прочее" in seg6:
        return "other"
    return None


def normalize_metric_name(value: Any) -> str:
    return normalize_text(value)


def normalize_sheet_name(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("ё", "е")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("ё", "е")
    text = re.sub(r"[^a-zа-я0-9]+", "", text)
    return text


def as_str(value: Any) -> str | None:
    if isinstance(value, str):
        text = value.strip()
        return text or None
    if value is None:
        return None
    return str(value)


def as_float(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        return to_float(value, None)
    return None


def to_float(value: Any, default: float | None) -> float | None:
    if isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return default

    text = value.strip().lower()
    if text in {"", "н/д", "na", "n/a", "#value!", "#div/0!", "-", "–"}:
        return default
    text = text.replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return default


def extract_year(iso_date: Any) -> int | None:
    if not isinstance(iso_date, str):
        return None
    parts = iso_date.split("-")
    if len(parts) != 3:
        return None
    if not parts[0].isdigit():
        return None
    return int(parts[0])


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9._-]+", "_", value)
    return cleaned[:200]


def write_csv(rows: list[dict[str, Any]], output_path: Path) -> None:
    all_keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                all_keys.append(key)

    with output_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=all_keys)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


if __name__ == "__main__":
    raise SystemExit(main())
